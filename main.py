import openpyxl as opx
import tkinter as tk
from openpyxl.styles import Font
from playsound import playsound
import time

def popup(x):
    def close_window():
        window.destroy()

    window = tk.Tk()
    window.title("Предупреждение!")
    window.geometry('300x150')
    l = tk.Label(window, text=x, font=("Arial Bold", 20), height=2)
    l.pack()
    btn = tk.Button(window, text="Понял, приму меры!", command=window.destroy)
    btn.pack()
    window.mainloop()





font = Font(name = 'Calibri', size = 11, color = 'FF0000')

while true:
    time.sleep(10)

    workbook = opx.load_workbook("D:\\test.xlsx", data_only=True)
    worksheet = workbook['Лист1']

    workbook2 = opx.load_workbook("D:\popup.xlsm", data_only=True)
    worksheet2 = workbook2['Лист2']

    for i in range(1, 5):
        for j in range(1, 4):
            worksheet.cell(row = i, column = j).value = worksheet2.cell(row = i, column = j).value

    workbook.save("D:\\test.xlsx")



    for i in range(1, 5):
        for j in range(2, 4):
            f = worksheet.cell(row = i, column = j).value
            d = float(f)
            if d >= 15:
                playsound("D:/Sound.mp3", block=False)
                name = worksheet.cell(row = i, column = 1).value
                worksheet.cell(row=i, column=j).font = font
                popup(name)
                workbook.save("D:\\test.xlsx")
            else:
                worksheet.cell(row=i, column=j).font = 'Calibri'
                workbook.save("D:\\test.xlsx")

workbook.save("D:\\test.xlsx")